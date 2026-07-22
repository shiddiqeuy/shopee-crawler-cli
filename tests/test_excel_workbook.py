"""Excel workbook tests."""

from datetime import UTC, datetime

from openpyxl import load_workbook

from shopee_cli.browser.models import BrowserMode
from shopee_cli.collectors.models import (
    SearchJob,
    SearchJobStatus,
    SearchResult,
    SearchSortMode,
)
from shopee_cli.exports.workbook_builder import build_workbook


def _job(collected_count: int = 2) -> SearchJob:
    return SearchJob(
        job_id="srch_test",
        keyword="kopi arabika",
        requested_limit=50,
        collected_count=collected_count,
        browser_mode=BrowserMode.MAIN,
        sort_mode=SearchSortMode.RELEVANCE,
        source_url="https://shopee.co.id/search?keyword=kopi+arabika",
        status=SearchJobStatus.COMPLETED,
        started_at=datetime(2026, 7, 22, 10, 0, tzinfo=UTC),
        finished_at=datetime(2026, 7, 22, 10, 5, tzinfo=UTC),
    )


def _results() -> list[SearchResult]:
    return [
        SearchResult(
            product_id="1",
            rank=1,
            product_name="Kopi A",
            product_url="https://shopee.co.id/Kopi-A-i.1.1",
            price_raw="Rp25.000",
            price_min=25000,
            price_max=25000,
            original_price=30000,
            discount_percentage=15,
            sold_raw="1,2RB terjual",
            sold_count=1200,
            rating=4.9,
            shop_name="Shop A",
            location="Jakarta",
            is_advertisement=True,
            is_mall=None,
            is_preferred=False,
            image_url="https://img.example/a.jpg",
            collected_at=datetime(2026, 7, 22, 10, 1, tzinfo=UTC),
        ),
        SearchResult(
            product_id="2",
            rank=2,
            product_name="Kopi B",
            product_url="https://shopee.co.id/Kopi-B-i.1.2",
            collected_at=datetime(2026, 7, 22, 10, 2, tzinfo=UTC),
        ),
    ]


def test_workbook_structure_metadata_and_reopen(tmp_path) -> None:
    """Workbook has two visible sheets, metadata, and can be reopened."""
    path = tmp_path / "report.xlsx"
    workbook = build_workbook(_job(), _results(), datetime(2026, 7, 22, 11, 0))
    workbook.save(path)

    reopened = load_workbook(path)

    assert reopened.sheetnames == ["Summary", "Search Results"]
    assert reopened.properties.title == "Shopee Market Research Report"
    assert reopened.properties.creator == "Muhammad Shiddiq Azis"


def test_workbook_writes_attribution_and_job_metadata() -> None:
    """Summary contains attribution and job metadata."""
    workbook = build_workbook(_job(), _results(), datetime(2026, 7, 22, 11, 0))
    summary = workbook["Summary"]

    values = [cell.value for row in summary.iter_rows() for cell in row]
    assert "srch_test" in values
    assert "Muhammad Shiddiq Azis" in values


def test_search_results_preserve_rank_order_and_rows() -> None:
    """Search Results writes all rows in stored rank order."""
    workbook = build_workbook(_job(), _results(), datetime(2026, 7, 22, 11, 0))
    sheet = workbook["Search Results"]

    assert sheet.max_row == 3
    assert [sheet.cell(row=row, column=1).value for row in (2, 3)] == [1, 2]


def test_search_results_formatting_and_hyperlinks() -> None:
    """Results sheet freezes header, filters, hyperlinks, and numeric formats."""
    workbook = build_workbook(_job(), _results(), datetime(2026, 7, 22, 11, 0))
    sheet = workbook["Search Results"]

    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    assert (
        sheet.cell(row=2, column=3).hyperlink.target
        == "https://shopee.co.id/Kopi-A-i.1.1"
    )
    assert sheet.cell(row=2, column=6).value == 25000
    assert "Rp" in sheet.cell(row=2, column=6).number_format
    assert sheet.cell(row=2, column=9).number_format == '0"%"'
    assert sheet.cell(row=2, column=15).value == "Yes"
    assert sheet.cell(row=2, column=16).value == "Unknown"


def test_zero_result_workbook_has_headers_only() -> None:
    """Valid no-results snapshots export without fake rows."""
    workbook = build_workbook(_job(collected_count=0), [], datetime(2026, 7, 22, 11, 0))
    sheet = workbook["Search Results"]

    assert sheet.max_row == 1
