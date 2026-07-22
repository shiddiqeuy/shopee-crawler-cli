"""Excel exporter orchestration tests."""

from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from shopee_cli.browser.models import BrowserMode
from shopee_cli.collectors.models import (
    SearchJobStatus,
    SearchRequest,
    SearchResult,
    SearchSortMode,
)
from shopee_cli.config.settings import Settings
from shopee_cli.database.connection import open_database
from shopee_cli.database.search_repository import SearchRepository
from shopee_cli.exports.excel_exporter import ExcelExporter
from shopee_cli.exports.exceptions import (
    ExportDataNotFoundError,
    ExportJobNotFoundError,
    ExportJobNotReadyError,
)


def _settings(tmp_path) -> Settings:
    return Settings(
        database_dir=tmp_path / "database",
        database_file_name="export.duckdb",
        export_path=tmp_path / "exports",
    )


def _request(keyword: str = "kopi") -> SearchRequest:
    return SearchRequest(
        keyword=keyword,
        requested_limit=10,
        browser_mode=BrowserMode.MAIN,
        sort_mode=SearchSortMode.RELEVANCE,
        max_scrolls=2,
        source_url=f"https://shopee.co.id/search?keyword={keyword}",
    )


def _result() -> SearchResult:
    return SearchResult(
        product_id="1",
        rank=1,
        product_name="Kopi",
        product_url="https://shopee.co.id/Kopi-i.1.1",
        collected_at=datetime.now(UTC),
    )


def test_exports_latest_snapshot_and_reopens_workbook(tmp_path) -> None:
    """Exporter writes latest eligible snapshot to configured export dir."""
    settings = _settings(tmp_path)
    repository = SearchRepository(settings)
    job = repository.create_job(_request())
    repository.save_results(job.job_id, [_result()])
    completed = repository.update_job_status(
        job.job_id,
        SearchJobStatus.COMPLETED,
        collected_count=1,
    )

    result = ExcelExporter(repository=repository, settings=settings).export()
    workbook = load_workbook(result.output_path)

    assert result.job.job_id == completed.job_id
    assert result.output_path.parent == settings.export_path
    assert workbook.sheetnames == ["Summary", "Search Results"]
    assert workbook["Search Results"].max_row == 2


def test_exports_by_job_id_and_keyword(tmp_path) -> None:
    """Exporter supports exact ID and latest keyword selection."""
    settings = _settings(tmp_path)
    repository = SearchRepository(settings)
    first = repository.create_job(_request("kopi"))
    repository.update_job_status(first.job_id, SearchJobStatus.COMPLETED)
    second = repository.create_job(_request("teh"))
    repository.update_job_status(second.job_id, SearchJobStatus.PARTIAL)

    exporter = ExcelExporter(repository=repository, settings=settings)

    assert exporter.export(job_id=first.job_id).job.job_id == first.job_id
    assert exporter.export(keyword="teh").job.job_id == second.job_id


def test_rejects_simultaneous_job_id_and_keyword(tmp_path) -> None:
    """Ambiguous selection is rejected."""
    exporter = ExcelExporter(settings=_settings(tmp_path))

    with pytest.raises(ExportJobNotFoundError):
        exporter.export(job_id="srch_x", keyword="kopi")


def test_reports_missing_job_and_not_ready_job(tmp_path) -> None:
    """Missing and failed jobs produce domain errors."""
    settings = _settings(tmp_path)
    repository = SearchRepository(settings)
    failed = repository.create_job(_request())
    repository.update_job_status(failed.job_id, SearchJobStatus.FAILED)
    exporter = ExcelExporter(repository=repository, settings=settings)

    with pytest.raises(ExportJobNotFoundError):
        exporter.export(job_id="srch_missing")
    with pytest.raises(ExportJobNotReadyError):
        exporter.export(job_id=failed.job_id)


def test_allows_zero_result_jobs(tmp_path) -> None:
    """Completed no-results snapshots can be exported."""
    settings = _settings(tmp_path)
    repository = SearchRepository(settings)
    job = repository.create_job(_request())
    repository.update_job_status(
        job.job_id, SearchJobStatus.COMPLETED, collected_count=0
    )

    result = ExcelExporter(repository=repository, settings=settings).export(
        job_id=job.job_id
    )

    assert load_workbook(result.output_path)["Search Results"].max_row == 1


def test_rejects_collected_count_mismatch(tmp_path) -> None:
    """Jobs claiming rows but containing none are rejected."""
    settings = _settings(tmp_path)
    repository = SearchRepository(settings)
    job = repository.create_job(_request())
    repository.update_job_status(
        job.job_id, SearchJobStatus.COMPLETED, collected_count=1
    )

    with pytest.raises(ExportDataNotFoundError):
        ExcelExporter(repository=repository, settings=settings).export(
            job_id=job.job_id
        )


def test_export_does_not_mutate_source_database_rows(tmp_path) -> None:
    """Export is read-only against search snapshot data."""
    settings = _settings(tmp_path)
    repository = SearchRepository(settings)
    job = repository.create_job(_request())
    repository.save_results(job.job_id, [_result()])
    repository.update_job_status(
        job.job_id, SearchJobStatus.COMPLETED, collected_count=1
    )
    before = _counts(settings)

    ExcelExporter(repository=repository, settings=settings).export(job_id=job.job_id)

    assert _counts(settings) == before


def _counts(settings: Settings) -> tuple[int, int]:
    with open_database(settings) as connection:
        jobs = connection.execute("SELECT COUNT(*) FROM search_jobs").fetchone()[0]
        results = connection.execute("SELECT COUNT(*) FROM search_results").fetchone()[
            0
        ]
    return jobs, results
