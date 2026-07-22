"""Summary sheet creation and statistics."""

from collections.abc import Iterable
from datetime import datetime
from statistics import median

from openpyxl.worksheet.worksheet import Worksheet

from shopee_cli.collectors.models import SearchJob, SearchResult
from shopee_cli.exports.models import SummaryStats
from shopee_cli.exports.styles import (
    apply_label_value_style,
    apply_section_header_style,
)


def calculate_summary_stats(results: list[SearchResult]) -> SummaryStats:
    """Calculate basic descriptive statistics from stored normalized fields."""
    prices = [result.price_min for result in results if result.price_min is not None]
    sold_counts = [
        result.sold_count for result in results if result.sold_count is not None
    ]
    ratings = [result.rating for result in results if result.rating is not None]

    return SummaryStats(
        total_products=len(results),
        products_with_price=len(prices),
        minimum_price=min(prices) if prices else None,
        maximum_price=max(prices) if prices else None,
        average_minimum_price=_average(prices),
        median_minimum_price=_median(prices),
        products_with_sold_data=len(sold_counts),
        average_sold_count=_average(sold_counts),
        median_sold_count=_median(sold_counts),
        products_with_rating=len(ratings),
        average_rating=_average(ratings),
        advertisement_count=sum(result.is_advertisement is True for result in results),
        mall_product_count=sum(result.is_mall is True for result in results),
        preferred_product_count=sum(result.is_preferred is True for result in results),
        unique_shops=_unique_non_empty(result.shop_name for result in results),
        unique_locations=_unique_non_empty(result.location for result in results),
        missing_product_name=sum(
            not _has_text(result.product_name) for result in results
        ),
        missing_product_url=sum(
            not _has_text(result.product_url) for result in results
        ),
        missing_price=sum(result.price_min is None for result in results),
        missing_sold_data=sum(result.sold_count is None for result in results),
        missing_rating=sum(result.rating is None for result in results),
        missing_shop_name=sum(not _has_text(result.shop_name) for result in results),
        missing_location=sum(not _has_text(result.location) for result in results),
        unknown_advertisement_status=sum(
            result.is_advertisement is None for result in results
        ),
    )


def create_summary_sheet(
    worksheet: Worksheet,
    job: SearchJob,
    results: list[SearchResult],
    exported_at: datetime,
) -> None:
    """Populate the Summary sheet."""
    stats = calculate_summary_stats(results)
    worksheet.title = "Summary"
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 70

    row = 1
    row = _write_section(worksheet, row, "Report Information")
    report_rows = [
        ("Report Title", "Shopee Market Research Report"),
        ("Keyword", job.keyword),
        ("Job ID", job.job_id),
        ("Search Status", job.status.value),
        ("Search URL", job.source_url),
        ("Browser Mode", job.browser_mode.value),
        ("Sort Mode", job.sort_mode.value),
        ("Requested Limit", job.requested_limit),
        ("Collected Products", job.collected_count),
        ("Search Started At", job.started_at),
        ("Search Finished At", job.finished_at),
        ("Exported At", exported_at),
    ]
    row = _write_label_values(worksheet, row, report_rows)

    row = _write_section(worksheet, row + 1, "Market Snapshot")
    market_rows = [
        ("Total Products", stats.total_products),
        ("Products With Price", stats.products_with_price),
        ("Minimum Price", stats.minimum_price),
        ("Maximum Price", stats.maximum_price),
        ("Average Minimum Price", _round_or_none(stats.average_minimum_price)),
        ("Median Minimum Price", _round_or_none(stats.median_minimum_price)),
        ("Products With Sold Data", stats.products_with_sold_data),
        ("Average Sold Count", _round_or_none(stats.average_sold_count)),
        ("Median Sold Count", _round_or_none(stats.median_sold_count)),
        ("Products With Rating", stats.products_with_rating),
        ("Average Rating", _round_or_none(stats.average_rating)),
        ("Advertisement Count", stats.advertisement_count),
        ("Mall Product Count", stats.mall_product_count),
        ("Preferred Product Count", stats.preferred_product_count),
        ("Unique Shops", stats.unique_shops),
        ("Unique Locations", stats.unique_locations),
    ]
    row = _write_label_values(worksheet, row, market_rows)

    row = _write_section(worksheet, row + 1, "Data Quality")
    quality_rows = [
        ("Missing Product Name", stats.missing_product_name),
        ("Missing Product URL", stats.missing_product_url),
        ("Missing Price", stats.missing_price),
        ("Missing Sold Data", stats.missing_sold_data),
        ("Missing Rating", stats.missing_rating),
        ("Missing Shop Name", stats.missing_shop_name),
        ("Missing Location", stats.missing_location),
        ("Unknown Advertisement Status", stats.unknown_advertisement_status),
    ]
    if stats.total_products == 0:
        quality_rows.append(
            ("Note", "No matching products were found during this snapshot.")
        )
    row = _write_label_values(worksheet, row, quality_rows)

    row = _write_section(worksheet, row + 1, "Attribution")
    _write_label_values(
        worksheet,
        row,
        [
            ("Created By", "Muhammad Shiddiq Azis"),
            (
                "Purpose",
                "To understand Shopee market condition through structured market "
                "research and data analysis.",
            ),
        ],
    )


def _write_section(worksheet: Worksheet, row: int, title: str) -> int:
    worksheet.cell(row=row, column=1, value=title)
    apply_section_header_style(worksheet[row])
    return row + 1


def _write_label_values(
    worksheet: Worksheet,
    row: int,
    values: list[tuple[str, object]],
) -> int:
    for label, value in values:
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=_excel_value(value))
        apply_label_value_style(worksheet[row])
        row += 1
    return row


def _average(values: Iterable[int | float]) -> float | None:
    numbers = list(values)
    return sum(numbers) / len(numbers) if numbers else None


def _median(values: Iterable[int | float]) -> float | None:
    numbers = list(values)
    return float(median(numbers)) if numbers else None


def _unique_non_empty(values: Iterable[str | None]) -> int:
    return len({value.strip() for value in values if value and value.strip()})


def _has_text(value: str | None) -> bool:
    return bool(value and value.strip())


def _round_or_none(value: float | None) -> float | None:
    return round(value, 2) if value is not None else None


def _excel_value(value: object) -> object:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value
