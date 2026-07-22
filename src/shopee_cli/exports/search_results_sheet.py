"""Search Results sheet creation."""

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from shopee_cli.collectors.models import SearchResult
from shopee_cli.exports.styles import (
    DATETIME_FORMAT,
    INTEGER_FORMAT,
    PERCENT_FORMAT,
    RATING_FORMAT,
    RUPIAH_FORMAT,
    THIN_BORDER,
    TOP_WRAP,
    apply_header_style,
)

HEADERS = [
    "Rank",
    "Product Name",
    "Product URL",
    "Product ID",
    "Price Displayed",
    "Price Minimum",
    "Price Maximum",
    "Original Price",
    "Discount %",
    "Sold Displayed",
    "Sold Count",
    "Rating",
    "Shop Name",
    "Location",
    "Advertisement",
    "Mall",
    "Preferred",
    "Image URL",
    "Collected At",
    "Job ID",
]

COLUMN_WIDTHS = {
    "A": 8,
    "B": 55,
    "C": 55,
    "D": 18,
    "E": 18,
    "F": 16,
    "G": 16,
    "H": 16,
    "I": 12,
    "J": 16,
    "K": 16,
    "L": 10,
    "M": 25,
    "N": 22,
    "O": 15,
    "P": 12,
    "Q": 12,
    "R": 50,
    "S": 22,
    "T": 36,
}


def create_search_results_sheet(
    worksheet: Worksheet,
    job_id: str,
    results: list[SearchResult],
) -> None:
    """Populate the Search Results sheet."""
    worksheet.title = "Search Results"
    worksheet.append(HEADERS)
    apply_header_style(worksheet[1])

    for result in results:
        worksheet.append(_result_row(result, job_id))
        row_number = worksheet.max_row
        url_cell = worksheet.cell(row=row_number, column=3)
        url_cell.hyperlink = result.product_url
        url_cell.style = "Hyperlink"
        image_cell = worksheet.cell(row=row_number, column=18)
        if result.image_url:
            image_cell.hyperlink = result.image_url
            image_cell.style = "Hyperlink"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _format_columns(worksheet)
    _add_table(worksheet)


def _result_row(result: SearchResult, job_id: str) -> list[object]:
    return [
        result.rank,
        result.product_name,
        result.product_url,
        result.product_id,
        result.price_raw,
        result.price_min,
        result.price_max,
        result.original_price,
        result.discount_percentage,
        result.sold_raw,
        result.sold_count,
        result.rating,
        result.shop_name,
        result.location,
        _boolean_label(result.is_advertisement),
        _boolean_label(result.is_mall),
        _boolean_label(result.is_preferred),
        result.image_url,
        result.collected_at.replace(tzinfo=None)
        if result.collected_at.tzinfo is not None
        else result.collected_at,
        job_id,
    ]


def _format_columns(worksheet: Worksheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TOP_WRAP
            cell.border = THIN_BORDER
    for column in ("F", "G", "H"):
        for cell in worksheet[column][1:]:
            cell.number_format = RUPIAH_FORMAT
    for column in ("I",):
        for cell in worksheet[column][1:]:
            cell.number_format = PERCENT_FORMAT
    for column in ("K",):
        for cell in worksheet[column][1:]:
            cell.number_format = INTEGER_FORMAT
    for column in ("L",):
        for cell in worksheet[column][1:]:
            cell.number_format = RATING_FORMAT
    for column in ("S",):
        for cell in worksheet[column][1:]:
            cell.number_format = DATETIME_FORMAT


def _add_table(worksheet: Worksheet) -> None:
    table = Table(displayName="SearchResults", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _boolean_label(value: bool | None) -> str:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return "Unknown"
