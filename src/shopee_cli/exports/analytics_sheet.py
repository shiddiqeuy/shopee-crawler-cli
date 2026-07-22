"""Analytics sheet creation."""

from openpyxl.worksheet.worksheet import Worksheet

from shopee_cli.analytics.engine import analyze_snapshot
from shopee_cli.analytics.models import AnalyticsReport
from shopee_cli.collectors.models import SearchJob, SearchResult
from shopee_cli.exports.styles import (
    THIN_BORDER,
    TOP_WRAP,
    apply_label_value_style,
    apply_section_header_style,
)


def create_analytics_sheet(
    worksheet: Worksheet,
    job: SearchJob,
    results: list[SearchResult],
) -> None:
    """Populate the Analytics sheet."""
    worksheet.title = "Analytics"
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 22
    report = analyze_snapshot(job, results)
    row = 1
    row = _write_section(worksheet, row, "Price Summary")
    row = _write_rows(
        worksheet,
        row,
        [
            ("Minimum Price", report.price.minimum),
            ("Maximum Price", report.price.maximum),
            ("Average Price", _round(report.price.average)),
            ("Median Price", _round(report.price.median)),
            ("Price Range", report.price.price_range),
            ("Q1", _round(report.price.q1)),
            ("Q3", _round(report.price.q3)),
            ("Products With Price", report.price.count),
            ("Missing Price", report.price.missing),
        ],
    )
    row = _write_section(worksheet, row + 1, "Sales Summary")
    row = _write_numeric_summary(worksheet, row, report, "sales", "Sold")
    row = _write_section(worksheet, row + 1, "Rating Summary")
    row = _write_numeric_summary(worksheet, row, report, "rating", "Rating")
    row = _write_section(worksheet, row + 1, "Seller Summary")
    row = _write_rows(
        worksheet,
        row,
        [
            ("Unique Shops", report.seller.unique_shops),
            ("Mall Shops", report.seller.mall_shops),
            ("Preferred Shops", report.seller.preferred_shops),
            ("Advertisement Products", report.seller.advertisement_products),
            ("Organic Products", report.seller.organic_products),
        ],
    )
    row = _write_section(worksheet, row + 1, "Location Summary")
    location_rows: list[tuple[str, object]] = [
        ("Unique Locations", report.location.unique_locations)
    ]
    location_rows.extend(report.location.top_locations)
    row = _write_rows(worksheet, row, location_rows)
    row = _write_section(worksheet, row + 1, "Product Distribution")
    row = _write_rows(
        worksheet,
        row,
        [
            ("Mall", report.distribution.mall),
            ("Preferred", report.distribution.preferred),
            ("Advertisement", report.distribution.advertisement),
            ("Unknown", report.distribution.unknown),
        ],
    )
    row = _write_section(worksheet, row + 1, "Data Quality")
    _write_rows(
        worksheet,
        row,
        [
            ("Missing Price", report.quality.missing_price),
            ("Missing Rating", report.quality.missing_rating),
            ("Missing Sold", report.quality.missing_sold),
            ("Missing Shop", report.quality.missing_shop),
            ("Missing Location", report.quality.missing_location),
            ("Duplicate Products Removed", report.quality.duplicate_products_removed),
        ],
    )
    for row_cells in worksheet.iter_rows():
        for cell in row_cells:
            cell.alignment = TOP_WRAP
            cell.border = THIN_BORDER


def _write_numeric_summary(
    worksheet: Worksheet,
    row: int,
    report: AnalyticsReport,
    attribute: str,
    label: str,
) -> int:
    summary = getattr(report, attribute)
    return _write_rows(
        worksheet,
        row,
        [
            (f"Minimum {label}", summary.minimum),
            (f"Maximum {label}", summary.maximum),
            (f"Average {label}", _round(summary.average)),
            (f"Median {label}", _round(summary.median)),
            (f"Products With {label} Data", summary.count),
            (f"Missing {label} Data", summary.missing),
        ],
    )


def _write_section(worksheet: Worksheet, row: int, title: str) -> int:
    worksheet.cell(row=row, column=1, value=title)
    apply_section_header_style(worksheet[row])
    return row + 1


def _write_rows(
    worksheet: Worksheet,
    row: int,
    values: list[tuple[str, object]],
) -> int:
    for label, value in values:
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=value)
        apply_label_value_style(worksheet[row])
        row += 1
    return row


def _round(value: float | None) -> float | None:
    return round(value, 2) if value is not None else None
