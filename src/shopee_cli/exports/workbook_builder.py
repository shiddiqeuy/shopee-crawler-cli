"""Workbook builder."""

from datetime import datetime

from openpyxl import Workbook

from shopee_cli.collectors.models import SearchJob, SearchResult
from shopee_cli.exports.analytics_sheet import create_analytics_sheet
from shopee_cli.exports.exceptions import WorkbookBuildError
from shopee_cli.exports.search_results_sheet import create_search_results_sheet
from shopee_cli.exports.summary_sheet import create_summary_sheet


def build_workbook(
    job: SearchJob,
    results: list[SearchResult],
    exported_at: datetime,
) -> Workbook:
    """Build the two-sheet Excel report workbook."""
    try:
        workbook = Workbook()
        summary = workbook.active
        create_summary_sheet(summary, job, results, exported_at)
        results_sheet = workbook.create_sheet("Search Results")
        create_search_results_sheet(results_sheet, job.job_id, results)
        analytics_sheet = workbook.create_sheet("Analytics")
        create_analytics_sheet(analytics_sheet, job, results)
        _apply_metadata(workbook)
        _validate_workbook(workbook)
    except Exception as error:
        msg = "Excel workbook could not be built."
        raise WorkbookBuildError(msg) from error
    return workbook


def _apply_metadata(workbook: Workbook) -> None:
    properties = workbook.properties
    properties.title = "Shopee Market Research Report"
    properties.creator = "Muhammad Shiddiq Azis"
    properties.description = (
        "Shopee market research report created to understand market condition "
        "through structured marketplace data analysis."
    )
    properties.subject = "Shopee Search Market Research"
    properties.keywords = (
        "Shopee, market research, keyword research, pricing, competition"
    )


def _validate_workbook(workbook: Workbook) -> None:
    visible_sheets = [
        sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"
    ]
    if [sheet.title for sheet in visible_sheets] != [
        "Summary",
        "Search Results",
        "Analytics",
    ]:
        msg = "Workbook must contain Summary, Search Results, and Analytics sheets."
        raise WorkbookBuildError(msg)
