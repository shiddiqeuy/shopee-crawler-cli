"""Centralized workbook styles."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color="1F4E78")
LABEL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)
RUPIAH_FORMAT = '"Rp" #,##0'
INTEGER_FORMAT = "#,##0"
PERCENT_FORMAT = '0"%"'
RATING_FORMAT = "0.##"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


def apply_header_style(row: tuple[object, ...]) -> None:
    """Apply table header styles."""
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_section_header_style(row: tuple[object, ...]) -> None:
    """Apply section header styles."""
    for cell in row[:2]:
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER


def apply_label_value_style(row: tuple[object, ...]) -> None:
    """Apply Summary label/value styles."""
    row[0].font = LABEL_FONT
    for cell in row[:2]:
        cell.alignment = TOP_WRAP
        cell.border = THIN_BORDER
